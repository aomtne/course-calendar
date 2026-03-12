#!/usr/bin/env python3
"""
Parse Excel (問卷系統_月曆顯示.xlsx) and regenerate index.html with updated RAW_DATA.
Used by GitHub Actions to auto-update the course calendar.

Course columns are dynamically detected from index 2 onward in each sheet's header row.
This means any new courses added to the Excel file are automatically included —
no code changes needed.
"""

import json
import re
import sys
import os
import glob
import openpyxl


def parse_date_cell(value, course_name):
    """
    Parse a cell value into (month, day, label).
    Patterns handled (in order):
      1. 'M/D (label)'         → month=M, day=D, label=label
      2. 'M/D ~ M/D'           → month=M1, day=D1, label='M1/D1~M2/D2'
      3. '20YY/M/D optional'   → month=M, day=D, label=optional or course_name
      4. 'M/D optional'        → month=M, day=D, label=optional or course_name
    """
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None

    # Pattern 1: M/D (label)
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*\(([^)]+)\)', s)
    if m:
        return int(m.group(1)), int(m.group(2)), m.group(3).strip()

    # Pattern 2: M/D ~ M/D range → take start date, use range as label
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*~\s*(\d{1,2})/(\d{1,2})', s)
    if m:
        sm, sd = int(m.group(1)), int(m.group(2))
        em, ed = int(m.group(3)), int(m.group(4))
        label = f"{sm}/{sd}~{em}/{ed}"
        return sm, sd, label

    # Pattern 3: 20YY/M/D optional-label
    m = re.match(r'^20\d{2}/(\d{1,2})/(\d{1,2})\s*(.*)', s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        label = m.group(3).strip() or course_name
        return month, day, label

    # Pattern 4: M/D optional-label
    m = re.match(r'^(\d{1,2})/(\d{1,2})\s*(.*)', s)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        label = m.group(3).strip() or course_name
        return month, day, label

    return None


def parse_name_map(wb):
    """Parse 姓名代碼 sheet to build code→name mapping."""
    name_map = {}
    if '姓名代碼' not in wb.sheetnames:
        return name_map
    ws = wb['姓名代碼']
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or row[1] is None:
            continue
        name = str(row[0]).strip()
        try:
            code = str(int(float(row[1])))
        except (ValueError, TypeError):
            code = str(row[1]).strip()
        if name and code:
            name_map[code] = name
    return name_map


def parse_excel(xlsx_path):
    """Parse the Excel file and return a list of RAW_DATA entries."""
    wb = openpyxl.load_workbook(xlsx_path)

    # Build code → name mapping
    name_map = parse_name_map(wb)

    # Collect data across all data sheets, keyed by (code, course_name)
    # Skip non-data sheets
    skip_sheets = {'代碼設定', '衝堂', '姓名代碼', '人員填寫狀況', '人員填寫情形'}
    all_entries = {}  # (code, course_name) → {month, day, label}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = list(rows[0])

        for row in rows[1:]:
            if not row or len(row) < 2:
                continue

            # Column 1 is the personal code
            raw_code = row[1]
            if raw_code is None:
                continue
            try:
                code = str(int(float(raw_code)))
            except (ValueError, TypeError):
                code = str(raw_code).strip()
            if not code:
                continue

            # Columns 2+ are course columns (dynamically detected)
            for col_idx in range(2, len(headers)):
                if col_idx >= len(row):
                    break
                course_name = headers[col_idx]
                if not course_name:
                    continue
                course_name = str(course_name).strip()
                # Clean up newlines in course names (e.g. from merged cells)
                course_name = course_name.replace('\n', '')

                cell_val = row[col_idx]
                result = parse_date_cell(cell_val, course_name)
                if result:
                    month, day, label = result
                    # Later sheets overwrite earlier ones for same (code, course)
                    all_entries[(code, course_name)] = {
                        "code": code,
                        "name": name_map.get(code, ""),
                        "courseName": course_name,
                        "month": month,
                        "day": day,
                        "label": label
                    }

    # Convert to sorted list
    data = list(all_entries.values())
    data.sort(key=lambda d: (d["code"], d["month"], d["day"], d["courseName"]))
    return data


def update_html(template_path, output_path, raw_data):
    """Replace RAW_DATA in the HTML template with new data."""
    with open(template_path, 'r', encoding='utf-8') as f:
        content = f.read()

    new_data_str = f"const RAW_DATA = {json.dumps(raw_data, ensure_ascii=False)};"

    # Replace the existing RAW_DATA declaration
    new_content = re.sub(
        r'const RAW_DATA = \[.*?\];',
        new_data_str,
        content,
        flags=re.DOTALL
    )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(new_content)

    return len(raw_data)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.dirname(script_dir)

    template_path = os.path.join(repo_root, 'template.html')
    output_path = os.path.join(repo_root, 'index.html')
    data_dir = os.path.join(repo_root, 'data')

    # Find Excel files in data/
    xlsx_files = glob.glob(os.path.join(data_dir, '*.xlsx'))
    if not xlsx_files:
        print("ERROR: No .xlsx files found in data/")
        sys.exit(1)

    # Use the most recently modified file
    latest_xlsx = max(xlsx_files, key=os.path.getmtime)
    print(f"Processing: {os.path.basename(latest_xlsx)}")

    # Parse Excel
    raw_data = parse_excel(latest_xlsx)
    print(f"Parsed {len(raw_data)} records")

    if not raw_data:
        print("WARNING: No data parsed from Excel. Aborting to avoid empty calendar.")
        sys.exit(1)

    # Update HTML
    count = update_html(template_path, output_path, raw_data)
    print(f"Updated index.html with {count} records")


if __name__ == '__main__':
    main()
